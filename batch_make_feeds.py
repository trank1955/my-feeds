#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Batch RSS generator + OPML per Feeder.
- Legge un Excel (due colonne: name, url)
- Per ogni riga genera un RSS .xml in output_feeds/
- Alla fine crea feeds.opml che punta ai .xml via GitHub RAW (o file:// se BASE_URL è vuoto)

Dipendenze: requests, beautifulsoup4, rfeed, openpyxl
Uso:
  python3 batch_make_feeds.py --excel feeds.xlsx --out ./output_feeds
"""

import argparse, os, re, html
from datetime import datetime, timezone
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
import rfeed
from openpyxl import load_workbook
import time

# ========= CONFIG =========
# Base URL RAW del repo: trank1955/my-feeds
BASE_URL = "https://raw.githubusercontent.com/trank1955/my-feeds/refs/heads/main/output_feeds"
UA = ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")
# ==========================

def slugify(text: str) -> str:
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-{2,}", "-", text).strip("-")
    return text or "feed"

def fetch_html(url: str, timeout: int = 25) -> BeautifulSoup:
    r = requests.get(url, headers={"User-Agent": UA}, timeout=timeout)
    r.raise_for_status()
    return BeautifulSoup(r.text, "html.parser")  # niente lxml

def _dz_fetch(url: str) -> BeautifulSoup:
    r = requests.get(url, headers={"User-Agent": UA}, timeout=20)
    r.raise_for_status()
    return BeautifulSoup(r.text, "html.parser")

def _dz_article_categories(article_url: str) -> list[str]:
    """Categorie/tassonomie dall'articolo Doppiozero.
    1) JSON-LD (articleSection/keywords)
    2) breadcrumb/tag fallback
    """
    try:
        soup = _dz_fetch(article_url)
    except Exception:
        return []

    cats: list[str] = []

    # --- (1) JSON-LD ---
    import json
    for s in soup.find_all("script", attrs={"type": "application/ld+json"}):
        try:
            data = json.loads(s.string or "")
        except Exception:
            continue
        # il JSON-LD può essere dict o list di dict
        blocks = data if isinstance(data, list) else [data]
        for b in blocks:
            if not isinstance(b, dict):
                continue
            # articleSection può essere str o list
            sec = b.get("articleSection")
            if isinstance(sec, str):
                cats.append(sec)
            elif isinstance(sec, list):
                cats.extend([str(x) for x in sec if x])
            # keywords a volte contiene categorie/temi
            kw = b.get("keywords")
            if isinstance(kw, str):
                # separa per virgola
                cats.extend([x.strip() for x in kw.split(",") if x.strip()])
            elif isinstance(kw, list):
                cats.extend([str(x) for x in kw if x])

    # --- (2) breadcrumb/tag fallback ---
    for a in soup.select(".breadcrumb a, nav.breadcrumb a"):
        t = a.get_text(" ", strip=True)
        if t:
            cats.append(t)

    for sel in [
        ".field--name-field-tags a",
        ".field--name-taxonomy-forums a",
        ".taxonomy-term a",
        ".meta a",
        ".node__meta a",
    ]:
        for a in soup.select(sel):
            t = a.get_text(" ", strip=True)
            if t:
                cats.append(t)

    # normalizza
    return [c.strip().lower() for c in cats if isinstance(c, str)]

def extract_items_generic(soup: BeautifulSoup, base_url: str):
    """
    Estrattore per pagine lista.
    - Preferenza Doppiozero: .view-content h2 a
    - Generico: <article> + fallback su h1/h2/h3 a
    """
    items = []

    # Doppiozero: teaser liste
    if "doppiozero.com" in base_url:
        for a in soup.select(".view-content h2 a"):
            href = a.get("href")
            title = a.get_text(" ", strip=True)
            if not href or not title:
                continue
            link = urljoin(base_url, href)
            p = a.find_parent("h2")
            desc_tag = p.find_next("p") if p else None
            desc = desc_tag.get_text(" ", strip=True) if desc_tag else None
            items.append({"title": title, "link": link, "desc": desc})

    # Generico: <article>
    if not items:
        for art in soup.select("article"):
            a = art.find("a", href=True)
            if not a:
                continue
            title = a.get_text(" ", strip=True) or a.get("title")
            link  = urljoin(base_url, a["href"])
            if not title or not link:
                continue
            p = art.find("p")
            desc = p.get_text(" ", strip=True) if p else None
            items.append({"title": title, "link": link, "desc": desc})

    # Fallback: titoli h1/h2/h3
    if not items:
        for h in soup.select("h1 a, h2 a, h3 a, h2, h3"):
            a = h if h.name == "a" else h.find("a", href=True)
            if not a or not a.get("href"):
                continue
            title = a.get_text(" ", strip=True) or a.get("title")
            link  = urljoin(base_url, a["href"])
            if not title or not link:
                continue
            desc_tag = (h if h.name != "a" else h.parent).find_next("p")
            desc = desc_tag.get_text(" ", strip=True) if desc_tag else None
            items.append({"title": title, "link": link, "desc": desc})

    # Dedup
    seen, out = set(), []
    for it in items:
        if it["link"] in seen:
            continue
        seen.add(it["link"])
        out.append(it)
    return out


def build_rss(name: str, base_url: str, items: list) -> str:
    now = datetime.now(timezone.utc)
    rss_items = []
    for it in items:
        rss_items.append(
            rfeed.Item(
                title=it["title"],
                link=it["link"],
                description=it.get("desc") or "",
                guid=rfeed.Guid(it["link"]),
                pubDate=now  # spesso la lista non espone date affidabili
            )
        )
    feed = rfeed.Feed(
        title=f"{name} (custom)",
        link=base_url,
        description=f"Feed generato automaticamente per {name}",
        language="it",
        lastBuildDate=now,
        items=rss_items
    )
    return feed.rss()

def write_opml_from_dir(dir_path: str, out_path: str, base_url: str | None) -> bool:
    files = [f for f in sorted(os.listdir(dir_path)) if f.lower().endswith(".xml")]
    if not files:
        return False
    lines = []
    for f in files:
        name = os.path.splitext(f)[0]
        text = html.escape(name)
        if base_url:
            xml_url = base_url.rstrip("/") + "/" + f
        else:
            full = os.path.abspath(os.path.join(dir_path, f))
            xml_url = "file://" + full
        lines.append(f'    <outline text="{text}" type="rss" xmlUrl="{html.escape(xml_url)}"/>')
    opml = """<?xml version="1.0" encoding="UTF-8"?>
<opml version="2.0">
  <head><title>Feeds export</title></head>
  <body>
{lines}
  </body>
</opml>
""".format(lines="\n".join(lines))
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(opml)
    return True

def read_excel_rows(xlsx_path: str):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [ (c.value.strip().lower() if isinstance(c.value, str) else c.value)
                for c in next(ws.iter_rows(min_row=1, max_row=1)) ]
    try:
        name_idx = headers.index("name")
        url_idx  = headers.index("url")
    except ValueError:
        raise SystemExit("L'Excel deve avere intestazioni: 'name' e 'url' sulla prima riga.")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: 
            continue
        name = (str(row[name_idx]).strip() if row[name_idx] else "")
        url  = (str(row[url_idx]).strip()  if row[url_idx]  else "")
        if name and url:
            rows.append((name, url))
    if not rows:
        raise SystemExit("Nessuna riga valida trovata (servono name e url).")
    return rows

def main():
    import argparse
    ap = argparse.ArgumentParser(description="Crea feed RSS da Excel e OPML finale.")
    ap.add_argument("--excel", required=True, help="File Excel (name,url)")
    ap.add_argument("--out", default="./output_feeds", help="Cartella output per i .xml")
    args = ap.parse_args()

    os.makedirs(args.out, exist_ok=True)
    rows = read_excel_rows(args.excel)

    ok = 0
    for name, url in rows:
        print(f"[+] {name} ← {url}")
        try:
            soup = fetch_html(url)
            items = extract_items_generic(soup, url)
            if not items:
                print("    \u26a0 Nessun articolo trovato (controlla che sia una pagina elenco).")
            xml = build_rss(name, url, items)
            fname = slugify(name) + ".xml"
            out_path = os.path.join(args.out, fname)
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(xml)
            print(f"    \u2714 {out_path} (items: {len(items)})")
            ok += 1
        except Exception as e:
            print(f"    \u2716 Errore: {e}")

    # OPML
    opml_path = os.path.join(args.out, "feeds.opml")
    base = BASE_URL if BASE_URL.strip() else None
    if write_opml_from_dir(args.out, opml_path, base):
        print(f"\n\u2714 OPML creato: {opml_path}")
        if base:
            print("   (Punta a GitHub RAW, pronto per Importa OPML in Feeder)")
        else:
            print("   (Usa file://; meglio pubblicare i .xml online e impostare BASE_URL)")
    else:
        print("\n\u26a0 Nessun .xml trovato per creare l'OPML.")

    print(f"\nFatto. Feed creati: {ok}")

if __name__ == "__main__":
    main()
